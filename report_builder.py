import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Constants ─────────────────────────────────────────────────────────────────

CLIENT_NAME_MAP = {
    "SeatGeek": "SeatGeek",
    "Gametime": "Gametime",
    "GoTickets": "GoTickets",
    "Mercury": "Mercury",
    "StubHub": "Stubhub",
    "Ticket Evolution": "Ticket Evolution",
    "TicketNetwork": "TicketNetwork",
    "TicketsNow": "TicketsNow",
    "TickPick": "TickPick",
    "Vivid Seats": "Vivid Seats",
}

NETWORK_ORDER = [
    "Gametime", "GoTickets", "Mercury", "Offsite", "SeatGeek",
    "Stubhub", "Ticket Evolution", "TicketNetwork", "TicketsNow",
    "TickPick", "Vivid Seats",
]

BUCKETS = ["Current", "1 to 30", "31 to 60", "61 to 90", "91 and Over"]

COMPANY_RENAMES = {
    "YS Tickets Spec": "YS Tickets",
    "YSA 2": "YSA",
    "YSA 3": "YSA",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def assign_bucket(days: int) -> str:
    if days <= 0:
        return "Current"
    elif days <= 30:
        return "1 to 30"
    elif days <= 60:
        return "31 to 60"
    elif days <= 90:
        return "61 to 90"
    else:
        return "91 and Over"


def load_and_filter(file, as_of_date: pd.Timestamp) -> pd.DataFrame:
    df = pd.read_excel(file)

    # Step 1: unpaid, not cancelled, balance > 0
    unpaid = df[
        (df["Paid"] == "No") &
        (df["IsCancelled"] == "No") &
        (df["Bal."] > 0)
    ].copy()

    # Step 2: exclude balances under $1
    unpaid = unpaid[unpaid["Bal."] >= 1].copy()

    # Step 3: remove duplicates — same Client + Ext Order #, only when Ext Order # is not blank
    has_ext = unpaid["Ext Order #"].notna() & (unpaid["Ext Order #"].astype(str).str.strip() != "")
    with_ext = unpaid[has_ext].drop_duplicates(subset=["Client", "Ext Order #"], keep="first")
    without_ext = unpaid[~has_ext]
    unpaid = pd.concat([with_ext, without_ext]).sort_index()

    # Step 4: compute aging days and bucket
    unpaid["days_out"] = (as_of_date - unpaid["Created"]).dt.days
    unpaid["bucket"] = unpaid["days_out"].apply(assign_bucket)

    # Step 5: tag rows that belong to "Other" network
    unpaid["is_other_network"] = ~unpaid["Client"].isin(CLIENT_NAME_MAP)

    return unpaid


def build_pivot(unpaid: pd.DataFrame):
    main = unpaid[~unpaid["is_other_network"]].copy()
    main["display_name"] = main["Client"].map(CLIENT_NAME_MAP)
    pivot = main.pivot_table(
        index="display_name", columns="bucket", values="Bal.",
        aggfunc="sum", fill_value=0,
    )
    other = unpaid[unpaid["is_other_network"]]
    other_by_bucket = other.groupby("bucket")["Bal."].sum()
    return pivot, other_by_bucket


def get_val(row_name, bkt, pivot, other_by_bucket) -> float:
    if row_name == "Offsite":
        return other_by_bucket.get(bkt, 0.0)
    if row_name in pivot.index and bkt in pivot.columns:
        return pivot.loc[row_name, bkt]
    return 0.0


# ── Excel builder ─────────────────────────────────────────────────────────────

def build_ar_aging_report(file, as_of_date: pd.Timestamp):
    unpaid = load_and_filter(file, as_of_date)
    pivot, other_by_bucket = build_pivot(unpaid)

    # Only include networks that have data
    active_rows = [
        rn for rn in NETWORK_ORDER
        if sum(get_val(rn, bkt, pivot, other_by_bucket) for bkt in BUCKETS) > 0
    ]

    wb = openpyxl.Workbook()
    _build_summary_sheet(wb, as_of_date, active_rows)
    _build_detail_sheet(wb, unpaid)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Build preview dataframe for Streamlit
    summary_rows = []
    for rn in active_rows:
        row = {"Network": rn}
        for bkt in BUCKETS:
            row[bkt] = get_val(rn, bkt, pivot, other_by_bucket)
        row["Total"] = sum(row[bkt] for bkt in BUCKETS)
        summary_rows.append(row)
    summary_df = pd.DataFrame(summary_rows)
    grand_total = summary_df["Total"].sum()

    return buf.read(), summary_df, grand_total, len(unpaid)


def _styles():
    thick = Side(style="medium", color="000000")
    return {
        "title": Font(name="Arial", bold=True, size=14),
        "subtitle": Font(name="Arial", bold=True, size=12),
        "date": Font(name="Arial", size=11),
        "header": Font(name="Arial", bold=True, size=11),
        "body": Font(name="Arial", size=11),
        "total": Font(name="Arial", bold=True, size=11),
        "center": Alignment(horizontal="center", vertical="center"),
        "thick": thick,
        "header_border": Border(top=thick, bottom=thick),
        "total_border": Border(top=thick, bottom=thick),
    }


def _build_summary_sheet(wb, as_of_date, active_rows):
    ws = wb.active
    ws.title = "AR Aging Summary"
    s = _styles()

    ws.column_dimensions["A"].width = 22
    for col in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 16

    # Titles
    titles = [
        ("Y&S Group (12 Entities)", s["title"]),
        ("A/R Aging Summary", s["subtitle"]),
        (f'As of {as_of_date.strftime("%B %-d, %Y")}', s["date"]),
    ]
    for r, (text, font) in enumerate(titles, 1):
        ws.merge_cells(f"A{r}:G{r}")
        c = ws.cell(row=r, column=1, value=text)
        c.font = font
        c.alignment = s["center"]

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 8

    # Column headers
    headers = ["Network", "Current", "1 to 30", "31 to 60", "61 to 90", "91 and Over", "Total"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=col_idx, value=h)
        c.font = s["header"]
        c.alignment = s["center"]
        c.border = s["header_border"]
    ws.row_dimensions[5].height = 18

    # Invoice Details column references (1-indexed):
    # C = Network, E = Amount, H = Aging
    detail_sheet = "'Invoice Details'"
    net_col  = f"{detail_sheet}!$C:$C"   # Network
    amt_col  = f"{detail_sheet}!$E:$E"   # Amount
    age_col  = f"{detail_sheet}!$H:$H"   # Aging

    # Data rows — SUMIFS(amount_col, network_col, network, aging_col, bucket)
    for i, row_name in enumerate(active_rows):
        r = 6 + i
        ws.row_dimensions[r].height = 16
        c = ws.cell(row=r, column=1, value=row_name)
        c.font = s["body"]
        c.alignment = s["center"]

        for col_idx, bkt in enumerate(BUCKETS, 2):
            formula = f'=SUMIFS({amt_col},{net_col},"{row_name}",{age_col},"{bkt}")'
            c = ws.cell(row=r, column=col_idx, value=formula)
            c.font = s["body"]
            c.alignment = s["center"]
            c.number_format = "$#,##0"

        # Row total = SUM of the five bucket cells
        first_bkt_col = get_column_letter(2)
        last_bkt_col  = get_column_letter(2 + len(BUCKETS) - 1)
        c = ws.cell(row=r, column=7, value=f"=SUM({first_bkt_col}{r}:{last_bkt_col}{r})")
        c.font = s["body"]
        c.alignment = s["center"]
        c.number_format = "$#,##0"

    # Total row — SUM of data rows per column
    total_row  = 6 + len(active_rows)
    first_data = 6
    last_data  = total_row - 1
    ws.row_dimensions[total_row].height = 18

    c = ws.cell(row=total_row, column=1, value="TOTAL")
    c.font = s["total"]
    c.alignment = s["center"]
    c.border = s["total_border"]

    for col_idx in range(2, 8):   # columns B–G
        col_letter = get_column_letter(col_idx)
        c = ws.cell(
            row=total_row, column=col_idx,
            value=f"=SUM({col_letter}{first_data}:{col_letter}{last_data})",
        )
        c.font = s["total"]
        c.alignment = s["center"]
        c.border = s["total_border"]
        c.number_format = "$#,##0"


def _build_detail_sheet(wb, unpaid: pd.DataFrame):
    ws = wb.create_sheet("Invoice Details")
    s = _styles()

    # Prepare source data
    source_df = unpaid.copy()
    source_df["Company"] = source_df["Company"].replace(COMPANY_RENAMES)

    # Map network: known clients get their display name, all others become "Offsite"
    source_df["Client_display"] = source_df["Client"].apply(
        lambda c: CLIENT_NAME_MAP.get(c, "Offsite")
    )

    # Column order: Broker, Invoice #, Network, Ext Order #, Amount, Status, Invoice Date, Aging
    source_cols_raw = ["Company", "Inv#", "Client_display", "Ext Order #", "Bal.", "Status", "Created", "bucket"]
    display_headers = {
        "Company":       "Broker",
        "Inv#":          "Invoice #",
        "Client_display":"Network",
        "Ext Order #":   "Ext Order #",
        "Bal.":          "Amount",
        "Status":        "Status",
        "Created":       "Invoice Date",
        "bucket":        "Aging",
    }
    col_widths = {
        "Broker": 22, "Invoice #": 14, "Network": 18, "Ext Order #": 20,
        "Amount": 16, "Status": 12, "Invoice Date": 20, "Aging": 14,
    }

    output_df = source_df[source_cols_raw].copy()

    # Write headers
    for col_idx, col in enumerate(source_cols_raw, 1):
        label = display_headers[col]
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font = s["header"]
        c.alignment = s["center"]
        c.border = s["header_border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[label]

    # Write data
    for row_idx, row in enumerate(output_df.itertuples(index=False), 2):
        for col_idx, (col, val) in enumerate(zip(source_cols_raw, row), 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = s["body"]
            c.alignment = s["center"]
            if col == "Bal.":
                c.number_format = "$#,##0.00"
            elif col == "Created":
                c.number_format = "MM/DD/YYYY"

    ws.freeze_panes = "A2"
